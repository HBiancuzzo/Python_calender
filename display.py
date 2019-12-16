import openpyxl as xl
from function import Calendar
import math


class Ui:
    space = 25
    instructions = '[L]eft|[R]ight|[M]onth|[D]ay|[E]vents of the month|[C]reate Event|[Q]uit'
    instructions_ce = '[L]eft|[R]ight|[D]ate|[H]our|[T]itle|[S]pan|[C]reate|[E]xit'
    len_L_R = len(instructions.split('|')[0]) + len(instructions.split('|')[1]) + 2

    @staticmethod
    def say_month(num_of_month):
        num_of_month_to_words = {
            1: 'January',
            2: 'February',
            3: 'March',
            4: 'April',
            5: 'May',
            6: 'June',
            7: 'July',
            8: 'August',
            9: 'September',
            10: 'October',
            11: 'November',
            12: 'December',
        }
        print('-' * 7 * (Ui.space + 2))
        line = ' ' + Ui.instructions.split('|')[0] + ' ' * (math.ceil(((Ui.space + 2) / 2) * 7) -
                                                            len(Ui.instructions.split('|')[0]) - math.ceil(
                    len(num_of_month_to_words[int(num_of_month)]) / 2)) + \
               num_of_month_to_words[int(num_of_month)]
        print(line + ' ' * ((Ui.space + 2) * 7 - len(line) - (len(Ui.instructions.split('|')[1])) - 1) +
              Ui.instructions.split('|')[1] + ' ')
        print('-' * (Ui.space + 2) * 7)

    @staticmethod
    def say_instructions(instructions):
        print()
        print('-' * (Ui.space + 2) * 7)
        num_ins = len(instructions.split('|')) - 2
        space_between = math.floor(
            ((Ui.space + 2) * 7) / (num_ins + 1) - (len(instructions) - Ui.len_L_R - num_ins) / (num_ins + 1))
        line = ' ' * space_between
        for ins in instructions.split('|'):
            if not ins == instructions.split('|')[0] and not ins == instructions.split('|')[1]:
                line += str(ins) + ' ' * space_between
        print(line)
        print('-' * (Ui.space + 2) * 7)

    @staticmethod
    def display_month(month):
        num_day_of_month = {
            1: 31, 3: 31, 5: 31, 7: 31, 8: 31, 10: 31, 12: 31,
            4: 30, 6: 30, 9: 30, 11: 30,
            2: 28
        }
        delay_from_month = {
            1: 1, 2: 4, 3: 4, 4: 0, 5: 2, 6: 5,
            7: 0, 8: 3, 9: 6, 10: 1, 11: 4, 12: 6,
        }
        day_of_week = {
            0: 'Monday',
            1: 'Tuesday',
            2: 'Wednesday',
            3: 'Thursday',
            4: 'Friday',
            5: 'Saturday',
            6: 'Sunday',
        }
        delay = delay_from_month[int(month)]
        current_day = 0

        # to name the 7 days of the week
        line = ''
        for num in range(7):
            line += '|' + ' ' * (math.floor(Ui.space / 2) - math.floor(len(day_of_week[num]) / 2)) + day_of_week[num] \
                    + ' ' * (math.ceil(Ui.space / 2) - math.ceil(len(day_of_week[num]) / 2)) + '|'
        print()
        print(line)

        # each month at maximum has 6 rows of weeks
        for num in range(6):
            array = ['', '', '', '', '', '']
            minimum = 1 + (7 * num) - delay
            maximum = 7 + (7 * num) + 1 - delay
            # writing for the 7 day of the week how many event are in each day and the box around
            for day in range(minimum, maximum):
                array[1] += '┌' + '-' * Ui.space + '┐'
                if 0 < day < num_day_of_month[int(month)] + 1:
                    array[2] += '|' + ' ' * (Ui.space - len(str(day)) - 1) + str(day) + ' ' + '|'
                else:
                    array[2] += '|' + ' ' * Ui.space + '|'
                month_day = month + '/' + str(day)
                event = Ui.display_date(month_day)[2]
                if event == 0:
                    array[3] += '|' + ' ' * Ui.space + '|'
                else:
                    array[3] += '|' + ' ' * (Ui.space - 17 - len(str(event))) + 'Nº of event/s: ' + str(
                        event) + ' ' * 2 + '|'
                array[4] += '|' + ' ' * Ui.space + '|'
                array[5] += '└' + '-' * Ui.space + '┘'
                current_day = day - 7

            # in case it has less rows of weeks
            if current_day > num_day_of_month[int(month)] - 1:
                break
            for i in range(len(array)):
                print(array[i])

    @staticmethod
    def display_date(date):
        num_of_month_to_words = {
            1: 'January',
            2: 'February',
            3: 'March',
            4: 'April',
            5: 'May',
            6: 'June',
            7: 'July',
            8: 'August',
            9: 'September',
            10: 'October',
            11: 'November',
            12: 'December',
        }
        db = xl.load_workbook('db.xlsx')
        db_sheet = db['Hoja1']
        count = 0
        # counts how many event happen in a day
        for row in range(2, db_sheet.max_row + 1):
            if date == db_sheet.cell(row, 2).value:
                count += 1
        return [num_of_month_to_words[int(date.split('/')[0])], date.split('/')[1], count]


def display_event_unite(date, time, m_or_d):
    num_of_month_to_words = {
        1: 'January',
        2: 'February',
        3: 'March',
        4: 'April',
        5: 'May',
        6: 'June',
        7: 'July',
        8: 'August',
        9: 'September',
        10: 'October',
        11: 'November',
        12: 'December',
    }
    month = int(date.split('/')[0])
    day = int(date.split('/')[1])
    title = Calendar.return_from_calendar(date, time)
    array = ['', '', '', '', '', '', 0]
    margin = 2

    if m_or_d == 'Month':
        string = f'{day} of {num_of_month_to_words[month]} at {time}'
    else:
        string = f'At {time}'

    len_date = len(string)
    len_title = len(str(title))

    # print the instant of the class calender

    if len_date >= len_title:
        space = len_date
    else:
        space = len_title
    space += (margin * 2)

    space_date = math.floor((space - len_date) / 2)
    space_title = math.floor((space - len_title) / 2)

    array[0] = '┌' + '-' * space + '┐'
    array[1] = '|' + ' ' * space_date + string + ' ' * (space - (space_date + len_date)) + '|'
    array[2] = '|' + ' ' * margin + '-' * (space - margin * 2) + ' ' * margin + '|'
    array[3] = '|' + ' ' * space_title + str(title) + ' ' * (space - (space_title + len_title)) + '|'
    array[4] = '|' + ' ' * space + '|'
    array[5] = '└' + '-' * space + '┘'
    array[6] = space + 2

    return array


def display_event_group(unite, m_or_d, day='1'):
    db = xl.load_workbook('db.xlsx')
    db_sheet = db['Hoja1']
    total_space = (Ui.space + 2) * 7
    count = 0
    array = ['', '', '', '', '', '']
    if m_or_d == 'Month':
        order_array = order_by_month(unite)
    elif m_or_d == 'Day':
        order_array = order_by_day(unite, day)
        print(' ' * math.floor((((Ui.space + 2) * 7) / 2) - (len(f'Day {day}') / 2)) + f'Day {day}')
        print(' ' * Ui.space + '-' * (Ui.space + 2) * 5)
    else:
        return

    # select each day that has an event of the month and print it

    if len(order_array) == 0:
        print('')
        print('')
        return 

    for index in order_array:
        date = db_sheet.cell(index, 2).value
        time = db_sheet.cell(index, 3).value
        array_day = display_event_unite(date, time, m_or_d)

        if count + array_day[6] <= total_space:
            count += array_day[6]
            for i in range(len(array)):
                array[i] += array_day[i]
        else:
            space = math.floor((total_space - count)/2)
            for info in array:
                info = ' ' * space + info + ' ' * (total_space - (space + count))
                print(info)

            count = array_day[6]
            for i in range(len(array)):
                array[i] = array_day[i]

    space = math.floor((total_space - count) / 2)
    for info in array:
        info = ' ' * space + info + ' ' * (total_space - (space + count))
        print(info)


def order_by_month(month):
    db = xl.load_workbook('db.xlsx')
    db_sheet = db['Hoja1']
    cell_index = []

    # select each day of the chosen month that has an event and add it to the array with row, day, hour, and minute
    for row in range(2, db_sheet.max_row + 1):
        data = []
        cell_month = int(db_sheet.cell(row, 2).value.split('/')[0])
        if int(month) == cell_month:
            data.append(row)
            data.append(int(db_sheet.cell(row, 2).value.split('/')[1]))
            data.append(int(db_sheet.cell(row, 3).value.split(':')[0]))
            data.append(int(db_sheet.cell(row, 3).value.split(':')[1]))
            cell_index.append(data)

    # sort the array in decreasing order without counting the row of the cell
    order_cell = []
    while not len(cell_index) == 0:
        prev = cell_index[0]
        for num in range(len(cell_index)):
            for i in range(1, 4):
                if cell_index[num][i] < prev[i]:
                    prev = cell_index[num]
                    break
                elif cell_index[num][i] > prev[i]:
                    break
        order_cell.append(prev[0])
        cell_index.remove(prev)
    return order_cell


def order_by_day(month, day):
    db = xl.load_workbook('db.xlsx')
    db_sheet = db['Hoja1']
    cell_index = []
    date = f'{month}/{day}'

    for row in range(2,  db_sheet.max_row + 1):
        data = []
        if db_sheet.cell(row, 2).value == date:
            data.append(row)
            data.append(db_sheet.cell(row, 3).value)
            cell_index.append(data)

    order_cell = []
    while not len(cell_index) == 0:
        prev = cell_index[0]
        for num in range(len(cell_index)):
            if cell_index[num][1] < prev[1]:
                prev = cell_index[num]
                break
            elif cell_index[num][1] > prev[1]:
                break
        order_cell.append(prev[0])
        cell_index.remove(prev)
    return order_cell


def display_create_event(date, time, title, span):
    # date, time, title, span
    array = [date, time, title, span]
    array_name = ['Date', 'Hour', 'Title', 'Span']

    print(' ' * math.floor((((Ui.space + 2) * 7) / 2) - (len('Create an event') / 2)) + 'Create an event')
    print(' ' * Ui.space + '-' * (Ui.space + 2) * 5)

    for i in range(len(array)):
        print()
        print(' ' * Ui.space + f'{array_name[i]}: {array[i]}')

