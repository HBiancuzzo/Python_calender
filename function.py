import openpyxl as xl
import math


class Calendar:

    db = xl.load_workbook('db.xlsx')
    db_sheet = db['Hoja1']

    def __init__(self, date, time, span):
        self.date = date
        self.time = time
        self.span = span

    def save_to_calendar(self, title):
        if self.is_free():
            max_cell = Calendar.db_sheet.max_row + 1
            cell = Calendar.db_sheet.cell(max_cell, 1)
            cell.value = title
            cell = Calendar.db_sheet.cell(max_cell, 2)
            cell.value = self.date
            cell = Calendar.db_sheet.cell(max_cell, 3)
            cell.value = self.time
            cell = Calendar.db_sheet.cell(max_cell, 4)
            cell.value = self.span
            Calendar.db.save('db.xlsx')
            return True
        else:
            return False

    @staticmethod
    def return_from_calendar(date, time):
        title = ''
        for row in range(2, Calendar.db_sheet.max_row + 1):
            cell = Calendar.db_sheet.cell(row, 2)
            if cell.value == date:
                cell = Calendar.db_sheet.cell(row, 3)
                if cell.value == time:
                    title = Calendar.db_sheet.cell(row, 1).value
        return title

    def is_free(self):
        # search for those that the same day
        for row in range(2, Calendar.db_sheet.max_row + 1):
            cell = Calendar.db_sheet.cell(row, 2)

            if cell.value == self.date:
                cell_span = Calendar.db_sheet.cell(row, 4).value
                # search for those that have the same hour than the object
                for c_minute in range(cell_span):
                    cell_hour = int(Calendar.db_sheet.cell(row, 3).value.split(':')[0])
                    cell_minute = int(Calendar.db_sheet.cell(row, 3).value.split(':')[1])
                    if c_minute + cell_minute >= 60:
                        cell_hour += 1
                        c_minute -= 60
                    # the s_hour is the max hour that span the event
                    s_hour = int(self.time.split(':')[0]) + math.floor((int(self.time.split(':')[1]) + self.span) / 60)
                    for hour in range(int(self.time.split(':')[0]), s_hour + 1):
                        if cell_hour == hour:
                            minimum = 0
                            maximum = 60

                            if cell_hour == int(self.time.split(':')[0]):  # in the min hour
                                minimum = int(self.time.split(':')[1])  # the minutes starts at the time of the event
                                if int(self.span) + int(self.time.split(':')[1]) <= 60:
                                    maximum = int(self.span) + int(self.time.split(':')[1])
                            elif cell_hour == s_hour:  # in the max hour the minutes could change so we have to change the search
                                maximum = int(self.time.split(':')[1]) + self.span - 60 * math.floor((int(self.time.split(':')[1]) + self.span) / 60)

                            for s_minute in range(minimum, maximum + 1):
                                if c_minute + cell_minute == s_minute:
                                    return False
        return True
